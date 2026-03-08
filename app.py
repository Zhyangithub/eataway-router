# ============================================================
# app.py — 完全自包含版本，不依赖 route_optimizer.py
# ============================================================
from flask import Flask, jsonify, render_template, request, send_file
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger
import json, os, io, urllib.parse, threading
from datetime import datetime
import pandas as pd
import requests as http_requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

app = Flask(__name__)

# ── 配置 ────────────────────────────────────────────────────
WAREHOUSE_COORD = "59.8542194,17.6650221"
DRIVERS        = ["Abbe", "Saman", "Sarkis", "Cornelia", "Pawlos"]

STATE_FILE  = "last_results.json"
PHONES_FILE  = "driver_phones.json"
EMAILS_FILE  = "driver_emails.json"
EMAIL_CONFIG_FILE = "email_config.json"

state = {
    "results": {},
    "generated_at": None,
    "schedule_hour": 7,
    "schedule_minute": 0,
    "running": False,
}
driver_phones  = {d: "" for d in DRIVERS}
driver_emails  = {
    d: os.environ.get(f"EMAIL_{d.upper()}", "")
    for d in DRIVERS
}
email_config   = {
    "sender":  os.environ.get("EMAIL_SENDER", "onboarding@resend.dev"),
    "api_key": os.environ.get("RESEND_API_KEY", ""),
}
scheduler = BackgroundScheduler()


# ── 持久化 ──────────────────────────────────────────────────
def load_state():
    if os.path.exists(STATE_FILE):
        with open(STATE_FILE, "r", encoding="utf-8") as f:
            saved = json.load(f)
            state["results"]         = saved.get("results", {})
            state["generated_at"]    = saved.get("generated_at")
            state["schedule_hour"]   = saved.get("schedule_hour", 7)
            state["schedule_minute"] = saved.get("schedule_minute", 0)
    if os.path.exists(PHONES_FILE):
        with open(PHONES_FILE, "r", encoding="utf-8") as f:
            driver_phones.update(json.load(f))
    if os.path.exists(EMAILS_FILE):
        with open(EMAILS_FILE, "r", encoding="utf-8") as f:
            driver_emails.update(json.load(f))
    if os.path.exists(EMAIL_CONFIG_FILE):
        with open(EMAIL_CONFIG_FILE, "r", encoding="utf-8") as f:
            saved_cfg = json.load(f)
            # 先从文件加载，再用环境变量覆盖
            if saved_cfg.get("sender"):
                email_config["sender"] = saved_cfg["sender"]
            if os.environ.get("EMAIL_SENDER"):
                email_config["sender"] = os.environ["EMAIL_SENDER"]
            if os.environ.get("RESEND_API_KEY"):
                email_config["api_key"] = os.environ["RESEND_API_KEY"]

def save_state():
    with open(STATE_FILE, "w", encoding="utf-8") as f:
        json.dump({
            "results":         state["results"],
            "generated_at":    state["generated_at"],
            "schedule_hour":   state["schedule_hour"],
            "schedule_minute": state["schedule_minute"],
        }, f, ensure_ascii=False, indent=2)

def save_phones():
    with open(PHONES_FILE, "w", encoding="utf-8") as f:
        json.dump(driver_phones, f, ensure_ascii=False, indent=2)

def save_emails():
    with open(EMAILS_FILE, "w", encoding="utf-8") as f:
        json.dump(driver_emails, f, ensure_ascii=False, indent=2)

def save_email_config():
    # api_key 仅通过环境变量管理，不写入文件
    with open(EMAIL_CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump({"sender": email_config.get("sender", "")},
                  f, ensure_ascii=False, indent=2)


# ── 核心逻辑 ─────────────────────────────────────────────────
def load_and_merge_data(driver_name):
    coords_file = 'coords.xlsx' if os.path.exists('coords.xlsx') else 'coords.csv'
    try:
        df_coords = pd.read_excel(coords_file, engine='openpyxl')
    except Exception as e:
        return [], f"读取坐标文件失败: {e}"

    header_row_idx = 0
    for i in range(min(15, len(df_coords))):
        if any('namn' in str(v).lower() for v in df_coords.iloc[i].tolist()):
            header_row_idx = i
            break
    df_coords.columns = df_coords.iloc[header_row_idx]
    df_coords = df_coords.iloc[header_row_idx+1:].reset_index(drop=True).iloc[:, :3]
    df_coords.columns = ['Namn', 'Latitude', 'Longitude']
    df_coords['match_name'] = df_coords['Namn'].astype(str).str.strip().str.lower()
    df_coords = df_coords.drop_duplicates(subset=['match_name'], keep='first')
    coord_dict = df_coords.set_index('match_name')[['Latitude', 'Longitude']].to_dict('index')

    routes_file = 'routes.xlsx' if os.path.exists('routes.xlsx') else 'routes.csv'
    try:
        df_routes = pd.read_excel(routes_file, engine='openpyxl')
    except Exception as e:
        return [], f"读取路线文件失败: {e}"

    driver_col = None
    name_row   = -1
    for col in df_routes.columns:
        if driver_name.lower() in str(col).lower():
            driver_col = col; break
        for r_idx in range(min(15, len(df_routes))):
            if driver_name.lower() in str(df_routes[col].iloc[r_idx]).lower():
                driver_col = col; name_row = r_idx; break
        if driver_col: break

    if driver_col is None:
        return [], f"找不到司机 {driver_name}"

    raw = (df_routes[driver_col].iloc[1:].tolist()
           if name_row == -1 else df_routes[driver_col].iloc[name_row+2:].tolist())

    matched, unmatched = [], []
    for store in raw:
        if pd.isna(store) or str(store).strip() in ("", "nan"):
            continue
        key = str(store).strip().lower()
        if key in coord_dict:
            matched.append({
                "name": str(store).strip(),
                "lat":  str(coord_dict[key]['Latitude']),
                "lng":  str(coord_dict[key]['Longitude']),
            })
        else:
            unmatched.append(str(store).strip())
    return matched, unmatched


def load_coord_dict():
    """Load name->coords mapping directly from coords source file."""
    coords_file = 'coords.xlsx' if os.path.exists('coords.xlsx') else 'coords.csv'
    try:
        df = pd.read_excel(coords_file, engine='openpyxl')
    except Exception:
        return {}
    header_row_idx = 0
    for i in range(min(15, len(df))):
        if any('namn' in str(v).lower() for v in df.iloc[i].tolist()):
            header_row_idx = i
            break
    df.columns = df.iloc[header_row_idx]
    df = df.iloc[header_row_idx+1:].reset_index(drop=True).iloc[:, :3]
    df.columns = ['Namn', 'Latitude', 'Longitude']
    df['match_name'] = df['Namn'].astype(str).str.strip().str.lower()
    df = df.drop_duplicates(subset=['match_name'], keep='first')
    return df.set_index('match_name')[['Latitude', 'Longitude']].to_dict('index')




def _distance_matrix_osrm(origins, destinations):
    """
    使用 OSRM 公共 API 计算距离/时间矩阵（完全免费，无需 API Key）。
    文档：http://project-osrm.org/docs/v5.24.0/api/#table-service
    返回 (time_matrix秒, dist_matrix米)，失败返回 (None, None)。
    """
    n_orig = len(origins)
    n_dest = len(destinations)

    # OSRM 格式：longitude,latitude（注意经纬顺序与 Google 相反）
    all_pts   = origins + destinations
    coords_str = ";".join(f"{float(p['lng'])},{float(p['lat'])}" for p in all_pts)
    src_indices = ";".join(str(i)         for i in range(n_orig))
    dst_indices = ";".join(str(n_orig + i) for i in range(n_dest))

    url = (
        f"https://router.project-osrm.org/table/v1/driving/{coords_str}"
        f"?sources={src_indices}&destinations={dst_indices}"
        f"&annotations=duration,distance"
    )

    print(f"[OSRM] Requesting {n_orig}×{n_dest} matrix ({n_orig * n_dest} elements)")
    try:
        resp = http_requests.get(url, timeout=30)
        data = resp.json()

        if data.get("code") != "Ok":
            print(f"[OSRM] ✗ code={data.get('code')} message={data.get('message','')}")
            return None, None

        time_matrix = data.get("durations")   # 秒，可能含 None（不可达）
        dist_matrix = data.get("distances")   # 米，可能含 None

        if not time_matrix:
            print("[OSRM] ✗ 返回数据中缺少 durations 字段")
            return None, None

        # distances 字段在部分 OSRM 版本中可能缺失，用零矩阵兜底
        if not dist_matrix:
            dist_matrix = [[0] * n_dest for _ in range(n_orig)]

        # 将不可达节点（None）替换为大值
        for i in range(n_orig):
            for j in range(n_dest):
                if time_matrix[i][j] is None:
                    time_matrix[i][j] = 999999
                if dist_matrix[i][j] is None:
                    dist_matrix[i][j] = 0

        print(f"[OSRM] ✓ {n_orig}×{n_dest} 矩阵成功")
        return time_matrix, dist_matrix

    except Exception as e:
        import traceback
        print(f"[OSRM] ✗ 异常: {e}\n{traceback.format_exc()}")
        return None, None


def _greedy_tsp_from(matrix, start=0):
    """
    贪心最近邻 TSP（Nearest Neighbor Heuristic）。
    作为 OR-Tools 失败时的保底方案。
    从 start 节点出发，每次选未访问中耗时最短的节点。
    返回完整访问顺序列表（不含起点）。
    """
    n = len(matrix)
    visited = [False] * n
    visited[start] = True
    order = []
    cur = start
    for _ in range(n - 1):
        best_j, best_t = -1, float("inf")
        for j in range(n):
            if not visited[j] and matrix[cur][j] < best_t:
                best_j, best_t = j, matrix[cur][j]
        if best_j == -1:
            break
        visited[best_j] = True
        order.append(best_j)
        cur = best_j
    for j in range(n):
        if not visited[j]:
            order.append(j)
    return order


def _ortools_tsp(matrix, start=0, locked_positions=None):
    """
    使用 Google OR-Tools 求解 TSP 全局最优路线。
    输入：行×列的秒数矩阵，start 为仓库节点索引。

    locked_positions: dict { 访问步数(int) → 矩阵节点索引(int) }
                      depot 出发时步数 = 0，第 1 站步数 = 1，以此类推。
                      例如 {2: 5, 4: 8} 表示：
                          第 2 步必须访问矩阵节点 5，
                          第 4 步必须访问矩阵节点 8。

                      实现方式：
                      1. CumulVar 维度约束：用 SetRange 钉死锁定节点的位置，
                         防止 GLS 元启发式在改进阶段移动它们。
                      2. 手动构建初始解：将锁定节点放在指定槽位，未锁定节点
                         按贪心最近邻填入剩余槽位。
                      3. SolveFromAssignment：直接从手动初始解开始 GLS 优化，
                         完全跳过 FirstSolutionStrategy（PATH_CHEAPEST_ARC /
                         LOCAL_CHEAPEST_INSERTION 等都无法可靠处理 CumulVar）。

    返回：(order, solved_with_locks)
        order: 访问顺序列表（不含起点 start），格式与 _greedy_tsp_from 相同。
        solved_with_locks: bool，True 表示含锁定约束求解成功，
                           False 表示降级（无约束或贪心）。
    """
    try:
        from ortools.constraint_solver import routing_enums_pb2
        from ortools.constraint_solver import pywrapcp
    except ImportError:
        print("[OR-TOOLS] ✗ ortools 未安装，回退到贪心算法")
        return _greedy_tsp_from(matrix, start), False

    n = len(matrix)
    if n <= 2:
        return [i for i in range(n) if i != start], (not locked_positions)

    def _extract_order(routing, manager, solution):
        """从 solution 中提取访问顺序列表（不含 start）。"""
        order = []
        index = routing.Start(0)
        while not routing.IsEnd(index):
            node = manager.IndexToNode(index)
            if node != start:
                order.append(node)
            index = solution.Value(routing.NextVar(index))
        return order

    def _build_base_model():
        """构建基础路由模型（不含锁定约束）。"""
        mgr = pywrapcp.RoutingIndexManager(n, 1, start)
        mdl = pywrapcp.RoutingModel(mgr)

        def time_callback(from_index, to_index):
            i = mgr.IndexToNode(from_index)
            j = mgr.IndexToNode(to_index)
            return int(matrix[i][j])

        transit_idx = mdl.RegisterTransitCallback(time_callback)
        mdl.SetArcCostEvaluatorOfAllVehicles(transit_idx)
        return mgr, mdl

    def _default_search_params():
        sp = pywrapcp.DefaultRoutingSearchParameters()
        sp.first_solution_strategy = (
            routing_enums_pb2.FirstSolutionStrategy.PATH_CHEAPEST_ARC
        )
        sp.local_search_metaheuristic = (
            routing_enums_pb2.LocalSearchMetaheuristic.GUIDED_LOCAL_SEARCH
        )
        sp.time_limit.seconds = 5
        sp.log_search = False
        return sp

    def _log_comparison(label, obj, locked_count=0):
        """对比 OR-Tools 解与贪心解并打印日志。"""
        greedy_order = _greedy_tsp_from(matrix, start)
        path = [start] + greedy_order
        greedy_sec = sum(matrix[path[i]][path[i + 1]] for i in range(len(path) - 1))
        greedy_sec += matrix[greedy_order[-1]][start]
        improvement = (greedy_sec - obj) / max(greedy_sec, 1) * 100
        lock_info = f" | 锁定 {locked_count}" if locked_count else ""
        print(f"[OR-TOOLS] ✓ {n} 节点{lock_info}{label} | "
              f"OR-Tools={round(obj/60)}min 贪心={round(greedy_sec/60)}min "
              f"节省={improvement:.1f}%")

    try:
        # ── 阶段 1：带锁定约束的全局优化 ──────────────────────
        if locked_positions:
            manager, routing = _build_base_model()

            # ─── 1a. 添加 CumulVar 维度约束 ─────────────────
            # 约束作用：在 GLS 改进阶段防止锁定节点被移动。
            # 不依赖 FirstSolutionStrategy——初始解由我们手动构建。
            def unit_transit(from_index, to_index):
                return 1

            unit_transit_idx = routing.RegisterTransitCallback(unit_transit)
            routing.AddDimension(
                unit_transit_idx,
                0,       # slack = 0（无松弛）
                n + 1,   # capacity（步数上限）
                True,    # fix_start_cumul_to_zero（depot 步数 = 0）
                "visit_order",
            )
            order_dim = routing.GetDimensionOrDie("visit_order")

            for step, node in locked_positions.items():
                routing_idx = manager.NodeToIndex(node)
                order_dim.CumulVar(routing_idx).SetRange(int(step), int(step))

            print(f"[OR-TOOLS] CumulVar 约束: {locked_positions}  (step → node)")

            # ─── 1b. 手动构建满足约束的初始解 ────────────────
            # 核心思路：将锁定节点放入指定槽位，
            # 剩余槽位用贪心最近邻（基于完整距离矩阵）填充未锁定节点。
            locked_node_set = set(locked_positions.values())
            unlocked_nodes  = [i for i in range(n)
                               if i != start and i not in locked_node_set]

            # 初始路线数组：route[0] = 第 1 站，route[1] = 第 2 站 …
            route = [None] * (n - 1)
            for step, node in locked_positions.items():
                route[step - 1] = node  # step 从 1 开始，数组从 0 开始

            # 贪心最近邻填充未锁定节点
            remaining = set(unlocked_nodes)
            for i in range(len(route)):
                if route[i] is not None:
                    continue
                # 前一个节点（用于计算距离）
                prev_node = start if i == 0 else route[i - 1]
                # 从 remaining 中选最近的
                best_node, best_cost = None, float('inf')
                for cand in remaining:
                    cost = matrix[prev_node][cand]
                    if cost < best_cost:
                        best_node, best_cost = cand, cost
                if best_node is not None:
                    route[i] = best_node
                    remaining.discard(best_node)

            # 安全检查：如果还有剩余节点（不应该发生），追加到末尾
            for leftover in remaining:
                for i in range(len(route)):
                    if route[i] is None:
                        route[i] = leftover
                        break

            print(f"[OR-TOOLS] 手动初始解: depot → {route[:5]}{'…' if len(route)>5 else ''} "
                  f"→ depot  ({len(route)} 站)")

            # ─── 1c. 从手动初始解开始优化 ────────────────────
            search_params = _default_search_params()
            routing.CloseModelWithParameters(search_params)
            initial_assignment = routing.ReadAssignmentFromRoutes([route], True)

            if initial_assignment:
                solution = routing.SolveFromAssignmentWithParameters(
                    initial_assignment, search_params,
                )
                if solution:
                    order = _extract_order(routing, manager, solution)
                    _log_comparison("", solution.ObjectiveValue(),
                                    len(locked_positions))
                    return order, True
                else:
                    print("[OR-TOOLS] ✗ SolveFromAssignment 返回 None（GLS 无法改进？）")
            else:
                print("[OR-TOOLS] ✗ ReadAssignmentFromRoutes 失败"
                      f"（路线可能不满足约束）: {route[:8]}…")

            print("[OR-TOOLS] ✗ 带锁定约束无解，降级到无约束全局优化…")

        # ── 阶段 2：无约束全局优化（降级） ────────────────────
        manager, routing = _build_base_model()
        search_params = _default_search_params()
        solution = routing.SolveWithParameters(search_params)

        if solution:
            order = _extract_order(routing, manager, solution)
            _log_comparison("（无约束降级）", solution.ObjectiveValue())
            return order, False

        print("[OR-TOOLS] ✗ 无约束也未找到解，回退到贪心算法")
        return _greedy_tsp_from(matrix, start), False

    except Exception as e:
        import traceback
        print(f"[OR-TOOLS] ✗ 异常，回退到贪心算法: {e}\n{traceback.format_exc()}")
        return _greedy_tsp_from(matrix, start), False


def _stats_from_matrices(full_order, time_matrix, dist_matrix):
    """
    直接从已有的时间/距离矩阵计算路线总统计，完全不调用任何 API。

    full_order: OR-Tools / 贪心返回的顺序列表（不含起点 0），
                路线为 0 → full_order[0] → ... → full_order[-1] → 0
    time_matrix: N×N 秒数矩阵
    dist_matrix: N×N 距离矩阵（米）

    返回 stats dict。
    """
    route = [0] + list(full_order)   # 仓库(0) → 所有门店
    total_sec  = 0
    total_dist = 0
    for i in range(len(route) - 1):
        a, b = route[i], route[i + 1]
        total_sec  += time_matrix[a][b]
        total_dist += dist_matrix[a][b]
    # 最后一站返回仓库
    total_sec  += time_matrix[route[-1]][0]
    total_dist += dist_matrix[route[-1]][0]

    hours = total_sec // 3600
    mins  = (total_sec % 3600) // 60
    dur_str = f"{hours}h {mins}min" if hours > 0 else f"{mins}min"
    print(f"[STATS-MATRIX] ✓ {dur_str} ({total_sec}s) {round(total_dist/1000,1)}km "
          f"[直接从矩阵计算，无额外 API 调用]")
    return {
        "duration_min":  round(total_sec / 60),
        "duration_sec":  total_sec,
        "distance_km":   round(total_dist / 1000, 1),
    }


def optimize_route(stores, locked_indices=None):
    """
    对门店列表进行路线优化（使用 OSRM + OR-Tools TSP）。

    locked_indices: set/list，stores 列表中需要锁定访问位置的索引（0-indexed）。
                    锁定门店会被转换为 OR-Tools 硬约束 {访问步数: 矩阵节点}，
                    在全局 N×N 距离矩阵中进行整体优化，未锁定门店自由调度。

    返回 (optimized_stores, stats_dict) 或 (None, error_string)。
    stats_dict 包含 duration_min, duration_sec, distance_km,
               以及 locks_honored (bool) 表示锁定约束是否被满足。
    """
    valid_stores = [s for s in stores if s.get('lat') and s.get('lng')
                    and str(s['lat']).strip() not in ('', 'nan', 'None')
                    and str(s['lng']).strip() not in ('', 'nan', 'None')]

    if not valid_stores:
        return None, "Inga butiker med giltiga koordinater"

    if len(valid_stores) == 1:
        return valid_stores, {"duration_min": 0, "duration_sec": 0,
                              "distance_km": 0.0, "locks_honored": True}

    print(f"[OPTIMIZE] stores={len(valid_stores)}")

    # ── 构建 locked_positions: {访问步数 → 矩阵节点索引} ──
    # stores 可能有坐标缺失而被过滤的项，需要建立 stores→valid_stores 索引映射。
    # 矩阵节点编号：0 = 仓库，1..n = valid_stores[0..n-1]。
    # 步数编号：depot = 0，第 1 站 = 1，第 2 站 = 2，…
    # 所以 valid_stores[k] 对应的矩阵节点 = k+1，步数也 = k+1。
    locked_positions = None  # dict {step(int): matrix_node(int)}
    if locked_indices:
        locked_set = set(locked_indices)
        # stores 原始索引 → valid_stores 索引
        valid_idx_map = {}
        vi = 0
        for orig_i, s in enumerate(stores):
            if (s.get('lat') and s.get('lng')
                    and str(s['lat']).strip() not in ('', 'nan', 'None')
                    and str(s['lng']).strip() not in ('', 'nan', 'None')):
                valid_idx_map[orig_i] = vi
                vi += 1

        # 构建 {step: node}
        # 前端发送的 stores 列表已是用户期望的顺序，
        # locked 的门店应该钉死在 valid_stores 空间中的对应位置。
        # valid_idx = valid_idx_map[orig_i]，矩阵节点 = valid_idx + 1，
        # 步数 = valid_idx + 1（depot=0 出发后第 1 站步数=1）。
        lp = {}
        for orig_i in sorted(locked_set):
            if orig_i not in valid_idx_map:
                continue
            valid_idx = valid_idx_map[orig_i]
            matrix_node = valid_idx + 1
            step = valid_idx + 1
            lp[step] = matrix_node
        if lp:
            locked_positions = lp
            print(f"[OPTIMIZE] locked_positions (step→node): {locked_positions} "
                  f"（stores 原始索引 {sorted(locked_set & set(valid_idx_map))}）")

    # ── OSRM 矩阵 + OR-Tools TSP ─────────────────────────────
    wh_lat, wh_lng = WAREHOUSE_COORD.split(',')
    warehouse  = {"lat": wh_lat, "lng": wh_lng}
    all_nodes  = [warehouse] + valid_stores

    time_m, dist_m = _distance_matrix_osrm(all_nodes, all_nodes)
    if time_m and len(time_m) == len(all_nodes):
        full_order, locks_honored = _ortools_tsp(
            time_m, start=0, locked_positions=locked_positions,
        )
        store_order = [idx - 1 for idx in full_order if idx > 0]
        optimized   = [valid_stores[i] for i in store_order]
        print(f"[OPTIMIZE] ✓ OR-Tools TSP order: {store_order} (locks_honored={locks_honored})")
        stats = _stats_from_matrices(full_order, time_m, dist_m)
        stats["locks_honored"] = locks_honored
        return optimized, stats

    # ── OSRM 失败时：Haversine + OR-Tools 保底（本地计算，零成本）──
    print("[OPTIMIZE] OSRM 失败，回退到 Haversine + OR-Tools（本地计算）…")
    import math

    def _haversine_sec(a, b, speed_kmh=35):
        R = 6371000
        lat1, lng1 = float(a["lat"]), float(a["lng"])
        lat2, lng2 = float(b["lat"]), float(b["lng"])
        dphi = math.radians(lat2 - lat1)
        dlam = math.radians(lng2 - lng1)
        h = math.sin(dphi/2)**2 + math.cos(math.radians(lat1))*math.cos(math.radians(lat2))*math.sin(dlam/2)**2
        dist_m = 2 * R * math.asin(math.sqrt(h)) * 1.35
        return int(dist_m / (speed_kmh / 3.6)), int(dist_m)

    n = len(all_nodes)
    fb_time = [[0]*n for _ in range(n)]
    fb_dist = [[0]*n for _ in range(n)]
    for i in range(n):
        for j in range(n):
            if i != j:
                fb_time[i][j], fb_dist[i][j] = _haversine_sec(all_nodes[i], all_nodes[j])

    full_order, locks_honored = _ortools_tsp(fb_time, start=0, locked_positions=locked_positions)
    store_order = [idx - 1 for idx in full_order if idx > 0]
    optimized   = [valid_stores[i] for i in store_order]
    stats = _stats_from_matrices(full_order, fb_time, fb_dist)
    stats["locks_honored"] = locks_honored
    print(f"[OPTIMIZE] Haversine fallback 完成: {store_order}")
    return optimized, stats


def get_route_stats(ordered_stores):
    """
    计算已排好序的路线的时间/距离统计（供 reorder 场景使用）。
    使用 OSRM 矩阵。
    Returns dict with duration_min, duration_sec, distance_km.
    """
    if not ordered_stores:
        return {"duration_min": 0, "duration_sec": 0, "distance_km": 0.0}

    wh_lat, wh_lng = WAREHOUSE_COORD.split(',')
    warehouse = {"lat": wh_lat, "lng": wh_lng}
    all_nodes = [warehouse] + list(ordered_stores)  # index 0=仓库, 1..n=门店

    print(f"[STATS] Fetching {len(all_nodes)}×{len(all_nodes)} matrix for route stats…")
    time_m, dist_m = _distance_matrix_osrm(all_nodes, all_nodes)

    if not time_m or len(time_m) != len(all_nodes):
        print(f"[STATS] ✗ Matrix fetch failed, returning None")
        return None

    # 路线顺序：仓库(0) → store[0](1) → store[1](2) → ... → store[n-1](n) → 仓库(0)
    route = list(range(len(all_nodes)))  # [0, 1, 2, ..., n]
    total_sec  = 0
    total_dist = 0
    for i in range(len(route) - 1):
        a, b = route[i], route[i + 1]
        total_sec  += time_m[a][b]
        total_dist += dist_m[a][b]
    total_sec  += time_m[route[-1]][0]   # 最后门店 → 仓库
    total_dist += dist_m[route[-1]][0]

    hours = total_sec // 3600
    mins  = (total_sec % 3600) // 60
    dur_str = f"{hours}h {mins}min" if hours > 0 else f"{mins}min"
    print(f"[STATS] ✓ {dur_str} ({total_sec}s) {round(total_dist/1000,1)}km")
    return {
        "duration_min":  round(total_sec / 60),
        "duration_sec":  total_sec,
        "distance_km":   round(total_dist / 1000, 1),
    }


def _gmaps_point(s):
    """Return coordinate string for warehouse, store name string for stores."""
    if s.get("is_warehouse"):
        return f"{s['lat']},{s['lng']}"
    return s["name"]


def generate_urls(optimized_stores):
    urls   = []
    wh_lat, wh_lng = WAREHOUSE_COORD.split(',')
    wh     = {"name": "Lager (Uppsala)", "lat": wh_lat, "lng": wh_lng, "is_warehouse": True}
    path   = [wh] + optimized_stores + [wh]
    for i in range(0, len(path) - 1, 10):
        chunk  = path[i: i+11]
        origin = chunk[0]; dest = chunk[-1]; wps = chunk[1:-1]
        wp_str = ""
        if wps:
            wp_str = "&waypoints=" + urllib.parse.quote(
                "|".join(_gmaps_point(s) for s in wps))
        urls.append(
            f"https://www.google.com/maps/dir/?api=1"
            f"&origin={urllib.parse.quote(_gmaps_point(origin))}"
            f"&destination={urllib.parse.quote(_gmaps_point(dest))}{wp_str}"
        )
    return urls


def run_all_drivers():
    results = {}
    print(f"[RUN_ALL] 开始为所有司机优化路线")

    for driver in DRIVERS:
        try:
            stores, unmatched = load_and_merge_data(driver)
            if not stores:
                err = unmatched if isinstance(unmatched, str) else "未匹配到任何门店"
                results[driver] = {"status": "error", "error": err}
                continue

            optimized, stats_or_err = optimize_route(stores)
            if not optimized:
                results[driver] = {"status": "error", "error": str(stats_or_err)}
                continue

            urls = generate_urls(optimized)

            dur_sec = stats_or_err.get('duration_sec', stats_or_err.get('duration_min', 0) * 60)
            hours   = dur_sec // 3600
            mins    = (dur_sec % 3600) // 60
            dur_str = f"{hours} h {mins} min" if hours > 0 else f"{mins} min"

            results[driver] = {
                "status":        "ok",
                "stores":        [s["name"] for s in optimized],
                "store_objects": optimized,
                "store_count":   len(optimized),
                "urls":          urls,
                "duration":      dur_str,
                "duration_sec":  dur_sec,
                "distance":      f"{stats_or_err['distance_km']} km",
                "unmatched":     unmatched if isinstance(unmatched, list) else [],
                "unmatched_count": len(unmatched) if isinstance(unmatched, list) else 0,
            }
            print(f"[RUN_ALL] {driver}: {dur_str} ({dur_sec}s) "
                  f"{stats_or_err['distance_km']}km")
        except Exception as e:
            results[driver] = {"status": "error", "error": str(e)}
    return results


# ── 后台任务 ─────────────────────────────────────────────────
def do_generate():
    # ★ running 已在 api_generate() 中设为 True（避免竞态）
    if not state["running"]:
        state["running"] = True
    try:
        state["results"]      = run_all_drivers()
        from datetime import timezone
        state["generated_at"] = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
        save_state()
    finally:
        state["running"] = False


def reschedule(hour, minute):
    if scheduler.get_job("daily_gen"):
        scheduler.remove_job("daily_gen")
    # ★ 必须指定时区，否则 CronTrigger 使用服务器默认时区（常为 UTC）
    #   导致用户设置 16:26 本地时间，实际在 UTC 16:26 才触发
    from zoneinfo import ZoneInfo
    tz = ZoneInfo("Europe/Stockholm")
    scheduler.add_job(do_generate, CronTrigger(hour=hour, minute=minute, timezone=tz),
                      id="daily_gen", replace_existing=True)
    print(f"[SCHEDULE] Armed daily job at {hour:02d}:{minute:02d} Europe/Stockholm")


# ── Flask 路由 ───────────────────────────────────────────────
@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/status")
def api_status():
    resp = jsonify({
        "results":         state["results"],
        "generated_at":    state["generated_at"],
        "schedule_hour":   state["schedule_hour"],
        "schedule_minute": state["schedule_minute"],
        "running":         state["running"],
        "drivers":         DRIVERS,
        "phones":          driver_phones,
        "emails":          driver_emails,
    })
    # 禁止浏览器/代理缓存路线结果
    resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    resp.headers["Pragma"]        = "no-cache"
    return resp


@app.route("/api/generate", methods=["POST"])
def api_generate():
    if state["running"]:
        return jsonify({"ok": False, "message": "Already running"}), 409
    # ★ FIX: 在启动线程之前设置 running=True，避免竞态条件：
    #   前端收到 ok 后立刻轮询 /api/status，如果线程还没来得及
    #   执行 do_generate() 里的 state["running"]=True，前端会
    #   看到 running=false + 旧结果，从而立刻停止轮询。
    state["running"] = True
    threading.Thread(target=do_generate, daemon=True).start()
    return jsonify({"ok": True})


@app.route("/api/schedule", methods=["POST"])
def api_schedule():
    data   = request.json
    hour   = int(data.get("hour",   7))
    minute = int(data.get("minute", 0))
    state["schedule_hour"]   = hour
    state["schedule_minute"] = minute
    reschedule(hour, minute)
    save_state()
    return jsonify({"ok": True, "hour": hour, "minute": minute})


@app.route("/api/phones", methods=["POST"])
def api_set_phones():
    for driver, number in request.json.items():
        if driver in driver_phones:
            driver_phones[driver] = str(number).strip()
    save_phones()
    return jsonify({"ok": True, "phones": driver_phones})


@app.route("/api/export")
def api_export():
    if not state["results"]:
        return jsonify({"error": "暂无结果"}), 400

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Route Links"

    hdr_fill   = PatternFill("solid", fgColor="1A1A2E")
    hdr_font   = Font(color="F5A623", bold=True, size=11)
    ok_fill    = PatternFill("solid", fgColor="0D2137")
    err_fill   = PatternFill("solid", fgColor="2D1A1A")
    wht_font   = Font(color="E0E0E0", size=10)
    center     = Alignment(horizontal="center",  vertical="center")
    left_wrap  = Alignment(horizontal="left",    vertical="center", wrap_text=True)
    thin       = Border(**{s: Side(style='thin', color='333355')
                           for s in ('left','right','top','bottom')})

    headers    = ["Chaufför","Butiker","Tid","Distans","Status",
                  "Segment 1","Segment 2","Segment 3"]
    widths     = [12, 8, 12, 12, 10, 60, 60, 60]

    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = hdr_fill; c.font = hdr_font
        c.alignment = center; c.border = thin
        ws.column_dimensions[c.column_letter].width = w
    ws.row_dimensions[1].height = 22

    for ri, driver in enumerate(DRIVERS, 2):
        r    = state["results"].get(driver, {})
        fill = ok_fill if r.get("status") == "ok" else err_fill
        ws.row_dimensions[ri].height = 30

        def mc(col, val, row=ri, f=fill):
            c = ws.cell(row=row, column=col, value=val)
            c.fill = f; c.font = wht_font; c.border = thin
            return c

        mc(1, driver).alignment = center
        mc(2, r.get("store_count", "—")).alignment = center
        mc(3, r.get("duration",    "—")).alignment = center
        mc(4, r.get("distance",    "—")).alignment = center
        mc(5, "Klar" if r.get("status")=="ok" else f"Fel: {r.get('error','')}").alignment = center
        for si in range(3):
            url = (r.get("urls") or [])[si] if si < len(r.get("urls") or []) else ""
            c = mc(6+si, url); c.alignment = left_wrap
            if url: c.font = Font(color="4A9FD4", size=9, underline="single")

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"rutter_{datetime.now().strftime('%Y%m%d')}.xlsx")



@app.route("/links/<driver_name>")
def driver_links(driver_name):
    r    = state["results"].get(driver_name)
    date = state.get("generated_at", "—")
    if not r or r.get("status") != "ok":
        return f"""<!DOCTYPE html><html><head><meta charset="UTF-8">
        <meta name="viewport" content="width=device-width,initial-scale=1">
        <title>{driver_name}</title></head>
        <body style="font-family:sans-serif;padding:2rem;background:#111;color:#fff">
        <h2>Inga rutter för {driver_name} ännu.</h2></body></html>""", 404

    urls          = r.get("urls", [])
    store_objects = r.get("store_objects", [
        {"name": s, "lat": "", "lng": ""} for s in r.get("stores", [])
    ])

    # Serialise store objects for injection into JS
    store_objects_json = json.dumps(store_objects, ensure_ascii=False)

    link_btns = "".join(
        f'<a href="{u}" id="mapbtn{i}" class="map-btn">🗺 Segment {i+1} — Öppna i Google Maps</a>'
        for i, u in enumerate(urls)
    )

    return f"""<!DOCTYPE html>
<html><head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width,initial-scale=1">
  <title>Körorder — {driver_name}</title>
  <style>
    *{{box-sizing:border-box;margin:0;padding:0}}
    body{{font-family:-apple-system,BlinkMacSystemFont,sans-serif;background:#0f111a;color:#e0e6f0;min-height:100vh}}
    .header{{background:#161b27;border-bottom:1px solid #1e2d45;padding:18px 20px}}
    .name{{font-size:26px;font-weight:800;color:#fff}}
    .meta{{font-size:13px;color:#6b7a99;margin-top:4px}}
    .stats{{display:grid;grid-template-columns:repeat(3,1fr);border-bottom:1px solid #1e2d45}}
    .stat{{padding:16px;text-align:center;border-right:1px solid #1e2d45}}
    .stat:last-child{{border-right:none}}
    .stat-val{{font-size:22px;font-weight:700;color:#f5a623;display:block}}
    .stat-lbl{{font-size:11px;color:#6b7a99;text-transform:uppercase;letter-spacing:.5px}}
    .section{{padding:16px 20px}}
    .section-title{{font-size:11px;color:#6b7a99;text-transform:uppercase;letter-spacing:1px;margin-bottom:12px}}
    .map-btn{{display:block;margin:10px 0;padding:14px 18px;background:#1a73e8;color:#fff;
              text-decoration:none;border-radius:8px;font-size:15px;font-weight:600;text-align:center}}
    .toggle{{background:none;border:1px solid #1e2d45;color:#6b7a99;padding:8px 14px;
             border-radius:6px;font-size:13px;cursor:pointer;margin-bottom:12px;width:100%}}

    /* ── Stop list ── */
    #stop-panel{{display:none;margin-top:4px}}
    .stop-list{{background:#161b27;border-radius:8px;overflow:hidden}}
    .stop-row{{display:flex;align-items:center;padding:10px 12px;
               border-bottom:1px solid #1e2d45;gap:8px;
               transition:background .15s;cursor:grab;user-select:none}}
    .stop-row:last-child{{border-bottom:none}}
    .stop-row.dragging{{opacity:.4;background:#0a0c14}}
    .stop-row.drag-over{{background:#1a2540;border-top:2px solid #1a73e8}}
    .stop-row.locked-row{{background:#1a1e10}}
    .stop-num{{color:#f5a623;font-weight:700;font-size:14px;min-width:26px;text-align:right}}
    .stop-name{{flex:1;font-size:14px;color:#e0e6f0}}
    .drag-handle{{color:#444;font-size:18px;cursor:grab;padding:0 4px}}
    .btn-up,.btn-dn{{background:none;border:1px solid #2a3550;color:#6b7a99;
                     border-radius:4px;font-size:14px;padding:3px 7px;cursor:pointer;line-height:1}}
    .btn-up:hover,.btn-dn:hover{{background:#1e2d45;color:#fff}}
    .btn-lock{{background:none;border:none;font-size:18px;cursor:pointer;padding:2px 4px;line-height:1}}
    .lock-hint{{font-size:11px;color:#6b7a99;text-align:center;padding:8px;font-style:italic}}

    /* ── Recalculate button ── */
    .recalc-bar{{padding:14px 20px;border-top:1px solid #1e2d45;background:#161b27;position:sticky;bottom:0}}
    .btn-recalc{{width:100%;padding:14px;background:#2d6a2d;color:#fff;border:none;
                 border-radius:8px;font-size:15px;font-weight:700;cursor:pointer}}
    .btn-recalc:hover{{background:#3a8a3a}}
    .btn-recalc:disabled{{background:#333;color:#666;cursor:not-allowed}}
    .recalc-status{{font-size:12px;color:#6b7a99;text-align:center;margin-top:6px;min-height:16px}}
    .locked-badge{{font-size:10px;background:#2a3510;color:#8bc34a;
                   border:1px solid #4a6420;border-radius:4px;padding:1px 5px;margin-left:6px}}
  </style>
</head>
<body>
  <div class="header">
    <div class="name">🚛 {driver_name}</div>
    <div class="meta">Genererad: {date}</div>
  </div>
  <div class="stats">
    <div class="stat"><span class="stat-val" id="s-count">{r.get("store_count","—")}</span><span class="stat-lbl">Butiker</span></div>
    <div class="stat"><span class="stat-val" id="s-dur">{r.get("duration","—")}</span><span class="stat-lbl">Est. tid</span></div>
    <div class="stat"><span class="stat-val" id="s-dist">{r.get("distance","—")}</span><span class="stat-lbl">Distans</span></div>
  </div>

  <div class="section">
    <div class="section-title">Navigationslänkar</div>
    <a href="/nav/{driver_name}" style="display:flex;align-items:center;gap:12px;padding:16px;background:#0d2a5a;color:#fff;text-decoration:none;border-radius:10px;font-size:15px;font-weight:700;margin-bottom:14px;border:2px solid #1a73e8">
      <span style="font-size:26px">📍</span>
      <div><div>Steg-för-steg navigation</div><div style="font-size:12px;font-weight:400;color:#7ab0f0;margin-top:2px">Ser butiksnamn i Google Maps ✓</div></div>
      <span style="margin-left:auto;font-size:20px">›</span>
    </a>
    <div class="section-title" style="margin-bottom:10px;margin-top:4px">Hela segmentet (klassisk)</div>
    <div id="map-btns">{link_btns}</div>

    <button class="toggle" id="toggle-btn"
      onclick="var p=document.getElementById('stop-panel');
               var open=p.style.display!='block';
               p.style.display=open?'block':'none';
               this.textContent=open?'▲ Dölj / redigera stopp':'▼ Visa / redigera stopp'">
      ▼ Visa / redigera stopp
    </button>

    <div id="stop-panel">
      <div class="lock-hint">🔒 Lås ett stopp för att hålla det kvar vid omräkning. Dra eller använd pilarna för att ändra ordning.</div>
      <div class="stop-list" id="stop-list"></div>
    </div>
  </div>

  <div class="recalc-bar" id="recalc-bar" style="display:none">
    <button class="btn-recalc" id="btn-recalc" onclick="recalculate()">
      🔄 Räkna om med låsta stopp
    </button>
    <div class="recalc-status" id="recalc-status"></div>
  </div>

<script>
// ── State ──────────────────────────────────────────────────────
const DRIVER = {json.dumps(driver_name)};
let stores   = {store_objects_json};
let locked   = new Set();   // Set of 0-indexed locked positions (in current order)
let dragSrcIdx = null;

// ── Render ─────────────────────────────────────────────────────
function render() {{
  const list = document.getElementById('stop-list');
  list.innerHTML = '';
  stores.forEach((s, i) => {{
    const isLocked = locked.has(i);
    const row = document.createElement('div');
    row.className = 'stop-row' + (isLocked ? ' locked-row' : '');
    row.draggable = true;
    row.dataset.idx = i;
    row.innerHTML = `
      <span class="drag-handle">⠿</span>
      <span class="stop-num">${{i+1}}</span>
      <span class="stop-name">${{s.name}}${{isLocked ? '<span class="locked-badge">🔒 LÅST</span>' : ''}}</span>
      <button class="btn-up" onclick="moveUp(${{i}})" ${{i===0?'disabled':''}}>▲</button>
      <button class="btn-dn" onclick="moveDown(${{i}})" ${{i===stores.length-1?'disabled':''}}>▼</button>
      <button class="btn-lock" onclick="toggleLock(${{i}})" title="${{isLocked?'Lås upp':'Lås position'}}">${{isLocked?'🔒':'🔓'}}</button>
    `;

    // Drag events
    row.addEventListener('dragstart', e => {{
      dragSrcIdx = i;
      setTimeout(() => row.classList.add('dragging'), 0);
    }});
    row.addEventListener('dragend', () => {{
      row.classList.remove('dragging');
      document.querySelectorAll('.stop-row').forEach(r => r.classList.remove('drag-over'));
    }});
    row.addEventListener('dragover', e => {{
      e.preventDefault();
      document.querySelectorAll('.stop-row').forEach(r => r.classList.remove('drag-over'));
      if (dragSrcIdx !== i) row.classList.add('drag-over');
    }});
    row.addEventListener('drop', e => {{
      e.preventDefault();
      if (dragSrcIdx !== null && dragSrcIdx !== i) {{
        moveToPos(dragSrcIdx, i);
      }}
    }});

    list.appendChild(row);
  }});

  // Show recalc bar only if any stop has been moved or locked
  updateRecalcBar();
}}

function updateRecalcBar() {{
  const bar = document.getElementById('recalc-bar');
  bar.style.display = 'block';   // always show once panel is used
}}

// ── Mutations ──────────────────────────────────────────────────
function moveUp(i) {{
  if (i === 0) return;
  swapStores(i, i-1);
}}
function moveDown(i) {{
  if (i === stores.length-1) return;
  swapStores(i, i+1);
}}
function swapStores(a, b) {{
  // Remap locked positions
  const newLocked = new Set();
  locked.forEach(p => {{
    if (p===a) newLocked.add(b);
    else if (p===b) newLocked.add(a);
    else newLocked.add(p);
  }});
  locked = newLocked;
  [stores[a], stores[b]] = [stores[b], stores[a]];
  render();
}}
function moveToPos(from, to) {{
  const item = stores.splice(from, 1)[0];
  stores.splice(to, 0, item);
  // Rebuild locked set: shift indices
  const arr = Array.from(locked);
  const newLocked = new Set();
  arr.forEach(p => {{
    if (p === from) {{ newLocked.add(to); return; }}
    let np = p;
    if (from < to) {{ if (p > from && p <= to) np = p - 1; }}
    else           {{ if (p >= to && p < from) np = p + 1; }}
    newLocked.add(np);
  }});
  locked = newLocked;
  render();
}}
function toggleLock(i) {{
  if (locked.has(i)) locked.delete(i);
  else locked.add(i);
  render();
}}

// ── Recalculate ────────────────────────────────────────────────
async function recalculate() {{
  const btn = document.getElementById('btn-recalc');
  const status = document.getElementById('recalc-status');
  btn.disabled = true;
  btn.textContent = '⏳ Räknar om…';
  status.textContent = '';

  try {{
    // Embed locked flag on each store object for the backend
    const payload = stores.map((s, i) => ({{...s, locked: locked.has(i)}}));
    const resp = await fetch(`/api/reorder/${{DRIVER}}`, {{
      method: 'POST',
      headers: {{'Content-Type': 'application/json'}},
      body: JSON.stringify({{ stores: payload }})
    }});
    const data = await resp.json();
    if (data.ok) {{
      stores = data.stores;
      locked = new Set();   // reset locks after successful recalc
      render();

      // Update stats
      document.getElementById('s-dur').textContent  = data.duration;
      document.getElementById('s-dist').textContent = data.distance;

      // Rebuild map buttons
      const btns = document.getElementById('map-btns');
      btns.innerHTML = data.urls.map((u,i) =>
        `<a href="${{u}}" class="map-btn">🗺 Segment ${{i+1}} — Öppna i Google Maps</a>`
      ).join('');

      btn.textContent = '✅ Klar! Räkna om igen';
      status.textContent = `Ny rutt: ${{data.duration}}, ${{data.distance}}`;
      if (data.warning) {{
        status.textContent += ` (⚠️ API-fel vid optimering: ${{data.warning}})`;
      }}
    }} else {{
      btn.textContent = '🔄 Räkna om med låsta stopp';
      status.textContent = '⚠️ Fel: ' + (data.error || data.msg || 'okänt fel');
    }}
  }} catch(e) {{
    btn.textContent = '🔄 Räkna om med låsta stopp';
    status.textContent = '⚠️ Nätverksfel: ' + e.message;
  }}
  btn.disabled = false;
}}

// ── Init ───────────────────────────────────────────────────────
render();
// Show recalc bar only after user opens the stop panel
document.getElementById('btn-recalc').closest('.recalc-bar').style.display = 'none';
document.getElementById('toggle-btn').addEventListener('click', () => {{
  const open = document.getElementById('stop-panel').style.display === 'block';
  document.getElementById('recalc-bar').style.display = open ? 'block' : 'none';
}});
</script>
</body></html>"""



def build_email_html(driver, r, base_url):
    date = state.get("generated_at", "—")
    page_url = f"{base_url}/links/{driver}"
    stores_html = "".join(
        f'<tr><td style="padding:6px 12px;color:#f5a623;width:30px">{i+1}</td>'
        f'<td style="padding:6px 12px">{s}</td></tr>'
        for i, s in enumerate(r.get("stores", []))
    )
    link_btns = "".join(
        f'<a href="{u}" style="display:block;margin:8px 0;padding:12px 16px;'
        f'background:#1a73e8;color:#fff;text-decoration:none;border-radius:6px;'
        f'font-size:14px;font-weight:600;text-align:center">Segment {i+1} — Oppna Google Maps</a>'
        for i, u in enumerate(r.get("urls", []))
    )
    return f"""<!DOCTYPE html>
<html><head><meta charset="UTF-8"></head>
<body style="font-family:-apple-system,sans-serif;background:#f5f5f5;margin:0;padding:20px">
  <div style="max-width:520px;margin:0 auto;background:#fff;border-radius:10px;overflow:hidden;box-shadow:0 2px 12px rgba(0,0,0,.1)">
    <div style="background:#0f111a;padding:20px 24px">
      <div style="font-size:22px;font-weight:800;color:#fff">Kororder — {driver}</div>
      <div style="font-size:13px;color:#6b7a99;margin-top:4px">Genererad: {date}</div>
    </div>
    <div style="display:grid;grid-template-columns:repeat(3,1fr);border-bottom:1px solid #eee">
      <div style="padding:16px;text-align:center;border-right:1px solid #eee">
        <div style="font-size:24px;font-weight:700;color:#f5a623">{r.get("store_count","—")}</div>
        <div style="font-size:11px;color:#999;text-transform:uppercase">Butiker</div>
      </div>
      <div style="padding:16px;text-align:center;border-right:1px solid #eee">
        <div style="font-size:24px;font-weight:700;color:#f5a623">{r.get("duration","—")}</div>
        <div style="font-size:11px;color:#999;text-transform:uppercase">Est. tid</div>
      </div>
      <div style="padding:16px;text-align:center">
        <div style="font-size:24px;font-weight:700;color:#f5a623">{r.get("distance","—")}</div>
        <div style="font-size:11px;color:#999;text-transform:uppercase">Distans</div>
      </div>
    </div>
    <div style="padding:20px 24px">
      <div style="font-size:11px;color:#999;text-transform:uppercase;letter-spacing:1px;margin-bottom:10px">Navigationslänkar</div>
      {link_btns}
      <div style="margin-top:16px">
        <a href="{page_url}" style="display:block;padding:12px 16px;background:#f0f7ff;color:#1a73e8;
           text-decoration:none;border-radius:6px;font-size:13px;text-align:center;border:1px solid #c8dffe">
          Oppna fullstandig ruttlank
        </a>
      </div>
      <div style="margin-top:20px">
        <div style="font-size:11px;color:#999;text-transform:uppercase;letter-spacing:1px;margin-bottom:8px">Korordning</div>
        <table style="width:100%;border-collapse:collapse;font-size:13px">{stores_html}</table>
      </div>
    </div>
    <div style="background:#f9f9f9;padding:14px 24px;border-top:1px solid #eee;font-size:12px;color:#999;text-align:center">
      RouteOps — Uppsala Warehouse
    </div>
  </div>
</body></html>"""


def send_email_to_driver(driver, r, base_url):
    to_addr = driver_emails.get(driver, "").strip()
    if not to_addr:
        return False, "Ingen e-postadress konfigurerad"
    api_key = email_config.get("api_key", "").strip()
    if not api_key:
        return False, "RESEND_API_KEY saknas — lägg till i Railway Variables"
    sender = email_config.get("sender", "onboarding@resend.dev").strip()
    try:
        html = build_email_html(driver, r, base_url)
        payload = {
            "from":    sender,
            "to":      [to_addr],
            "subject": f"Kororder {driver} — {state.get('generated_at','')}",
            "html":    html,
        }
        resp = http_requests.post(
            "https://api.resend.com/emails",
            headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
            json=payload,
            timeout=15
        )
        if resp.status_code in (200, 201):
            return True, "Skickat"
        else:
            err = resp.text
            print(f"[EMAIL ERROR] {driver}: {resp.status_code} {err}")
            return False, f"Resend API fel {resp.status_code}: {err}"
    except Exception as e:
        import traceback
        print(f"[EMAIL ERROR] {driver}: {traceback.format_exc()}")
        return False, str(e)

@app.route("/nav/<driver_name>")
def driver_nav(driver_name):
    r    = state["results"].get(driver_name)
    date = state.get("generated_at", "—")
    if not r or r.get("status") != "ok":
        return f"""<!DOCTYPE html><html><head><meta charset="UTF-8">
        <meta name="viewport" content="width=device-width,initial-scale=1">
        <title>{driver_name}</title></head>
        <body style="font-family:sans-serif;padding:2rem;background:#111;color:#fff">
        <h2>Inga rutter för {driver_name} ännu.</h2></body></html>""", 404

    store_objects = r.get("store_objects", [
        {"name": s, "lat": "", "lng": ""} for s in r.get("stores", [])
    ])
    wh_lat, wh_lng = WAREHOUSE_COORD.split(',')
    all_stops = [{"name": "🏭 Lager (Uppsala)", "lat": wh_lat, "lng": wh_lng, "is_warehouse": True}]
    all_stops += store_objects
    all_stops += [{"name": "🏭 Lager (Uppsala)", "lat": wh_lat, "lng": wh_lng, "is_warehouse": True}]
    stops_json  = json.dumps(all_stops, ensure_ascii=False)
    driver_json = json.dumps(driver_name)

    return f"""<!DOCTYPE html>
<html><head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width,initial-scale=1,maximum-scale=1">
  <title>Navigering — {driver_name}</title>
  <style>
    *{{box-sizing:border-box;margin:0;padding:0}}
    body{{font-family:-apple-system,BlinkMacSystemFont,sans-serif;background:#0f111a;color:#e0e6f0;min-height:100vh;display:flex;flex-direction:column}}
    .header{{background:#161b27;border-bottom:1px solid #1e2d45;padding:14px 16px;display:flex;align-items:center;gap:10px}}
    .back-btn{{color:#6b7a99;font-size:22px;text-decoration:none;line-height:1}}
    .header-info .driver{{font-size:18px;font-weight:800;color:#fff}}
    .header-info .meta{{font-size:12px;color:#6b7a99;margin-top:2px}}
    .progress-bar{{height:5px;background:#1e2d45}}
    .progress-fill{{height:100%;background:#f5a623;transition:width .4s ease}}
    .main{{flex:1;padding:16px;display:flex;flex-direction:column;gap:12px}}
    .status-label{{font-size:11px;color:#6b7a99;text-transform:uppercase;letter-spacing:1px;font-weight:600}}
    .stop-card{{background:#161b27;border-radius:14px;border:2px solid #1e2d45;padding:18px}}
    .stop-card.current-card{{border-color:#f5a623}}
    .stop-card.next-card{{opacity:.8}}
    .stop-badge{{font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:.8px;margin-bottom:6px}}
    .current-badge{{color:#f5a623}}
    .next-badge{{color:#6b7a99}}
    .stop-number{{font-size:13px;color:#6b7a99;margin-bottom:4px}}
    .stop-name{{font-size:26px;font-weight:800;color:#fff;line-height:1.2;word-break:break-word}}
    .stop-name.warehouse{{font-size:20px;color:#8bc34a}}
    .btn-nav{{display:flex;align-items:center;justify-content:center;gap:10px;width:100%;padding:18px;border:none;border-radius:12px;background:#1a73e8;color:#fff;font-size:17px;font-weight:700;cursor:pointer;text-decoration:none}}
    .btn-nav:active{{background:#1558b0}}
    .btn-arrived{{display:flex;align-items:center;justify-content:center;gap:10px;width:100%;padding:16px;border:none;border-radius:12px;background:#2d6a2d;color:#fff;font-size:16px;font-weight:700;cursor:pointer}}
    .btn-arrived:active{{background:#1e4a1e}}
    .btn-arrived.final{{background:#5a2d7a}}
    .stop-list-section{{background:#161b27;border-radius:12px;overflow:hidden}}
    .stop-list-title{{font-size:11px;color:#6b7a99;text-transform:uppercase;letter-spacing:1px;padding:12px 14px 8px;border-bottom:1px solid #1e2d45}}
    .stop-item{{display:flex;align-items:center;padding:10px 14px;border-bottom:1px solid #1a2030;gap:10px}}
    .stop-item:last-child{{border-bottom:none}}
    .stop-item.done{{opacity:.35}}
    .stop-item.active-item{{background:#1a2540}}
    .item-num{{font-size:12px;color:#f5a623;font-weight:700;min-width:22px;text-align:right}}
    .item-num.done-num{{color:#2d6a2d}}
    .item-name{{font-size:14px;color:#e0e6f0;flex:1}}
    .done-screen{{display:none;flex:1;flex-direction:column;align-items:center;justify-content:center;text-align:center;padding:32px 24px;gap:16px}}
    .done-icon{{font-size:72px}}
    .done-title{{font-size:28px;font-weight:800;color:#8bc34a}}
    .done-sub{{font-size:15px;color:#6b7a99}}
    .btn-reset{{padding:14px 28px;background:#333;color:#aaa;border:none;border-radius:8px;font-size:14px;cursor:pointer;margin-top:8px}}
  </style>
</head>
<body>
  <div class="header">
    <a href="/links/{driver_name}" class="back-btn">←</a>
    <div class="header-info">
      <div class="driver">🚛 {driver_name}</div>
      <div class="meta" id="header-meta">Laddar…</div>
    </div>
  </div>
  <div class="progress-bar"><div class="progress-fill" id="prog-fill" style="width:0%"></div></div>
  <div class="main" id="main-view">
    <div>
      <div class="status-label" style="margin-bottom:8px">Nuvarande destination</div>
      <div class="stop-card current-card">
        <div class="stop-badge current-badge">▶ Navigerar till</div>
        <div class="stop-number" id="cur-num"></div>
        <div class="stop-name" id="cur-name"></div>
      </div>
    </div>
    <a id="btn-nav" class="btn-nav" href="#" onclick="openNav(event)">
      <span style="font-size:22px">🗺</span>
      <span id="nav-btn-text">Öppna Google Maps navigation</span>
    </a>
    <button class="btn-arrived" id="btn-arrived" onclick="markArrived()">✅ Framme — nästa stopp</button>
    <div>
      <div class="status-label" style="margin-bottom:8px">Nästa stopp</div>
      <div class="stop-card next-card" id="next-card">
        <div class="stop-badge next-badge">Därefter</div>
        <div class="stop-number" id="nxt-num"></div>
        <div class="stop-name" id="nxt-name"></div>
      </div>
    </div>
    <div class="stop-list-section">
      <div class="stop-list-title">Alla stopp — {r.get("store_count","?")} butiker</div>
      <div id="stop-list-items"></div>
    </div>
  </div>
  <div class="done-screen" id="done-screen">
    <div class="done-icon">🎉</div>
    <div class="done-title">Rutten klar!</div>
    <div class="done-sub">Alla stopp besökta.<br>Bra jobbat, {driver_name}!</div>
    <button class="btn-reset" onclick="resetRoute()">↺ Börja om från lager</button>
  </div>
<script>
const DRIVER = {driver_json};
const STOPS  = {stops_json};
const KEY    = 'nav_v2_' + DRIVER;
let curIdx = parseInt(localStorage.getItem(KEY) || '1', 10);
if (isNaN(curIdx) || curIdx < 1 || curIdx >= STOPS.length) curIdx = 1;
function save() {{ localStorage.setItem(KEY, curIdx); }}
function openNav(e) {{
  e.preventDefault();
  const s = STOPS[curIdx];
  let dest = (s.lat && s.lng) ? encodeURIComponent(s.lat + ',' + s.lng) : encodeURIComponent(s.name);
  window.open('https://www.google.com/maps/dir/?api=1&destination=' + dest + '&travelmode=driving', '_blank');
}}
function markArrived() {{ curIdx++; save(); render(); window.scrollTo(0,0); }}
function resetRoute() {{ curIdx = 1; save(); render(); }}
function render() {{
  if (curIdx >= STOPS.length) {{
    document.getElementById('main-view').style.display   = 'none';
    document.getElementById('done-screen').style.display = 'flex';
    document.getElementById('prog-fill').style.width = '100%';
    document.getElementById('header-meta').textContent = 'Rutten klar! 🎉';
    return;
  }}
  document.getElementById('main-view').style.display   = 'flex';
  document.getElementById('done-screen').style.display = 'none';
  const pct = Math.round(((curIdx-1)/(STOPS.length-1))*100);
  document.getElementById('prog-fill').style.width = pct + '%';
  const stopsLeft = STOPS.length - 1 - curIdx;
  document.getElementById('header-meta').textContent = 'Stopp ' + curIdx + ' av ' + (STOPS.length-1) + ' · ' + stopsLeft + ' kvar';
  const cur = STOPS[curIdx];
  document.getElementById('cur-num').textContent = (curIdx===STOPS.length-1) ? 'Slutdestination' : 'Stopp ' + curIdx;
  const curNameEl = document.getElementById('cur-name');
  curNameEl.textContent = cur.name;
  curNameEl.className = 'stop-name' + (cur.is_warehouse ? ' warehouse' : '');
  document.getElementById('nav-btn-text').textContent = 'Navigera till ' + cur.name;
  const arrivedBtn = document.getElementById('btn-arrived');
  if (curIdx === STOPS.length-1) {{
    arrivedBtn.textContent = '🏁 Framme på lagret — avsluta rutten';
    arrivedBtn.className = 'btn-arrived final';
  }} else {{
    arrivedBtn.textContent = '✅ Framme — nästa stopp';
    arrivedBtn.className = 'btn-arrived';
  }}
  const nextIdx = curIdx + 1;
  const nextCard = document.getElementById('next-card');
  if (nextIdx < STOPS.length) {{
    document.getElementById('nxt-num').textContent = (nextIdx===STOPS.length-1) ? 'Slutdestination' : 'Stopp ' + nextIdx;
    document.getElementById('nxt-name').textContent = STOPS[nextIdx].name;
    nextCard.style.display = 'block';
  }} else {{ nextCard.style.display = 'none'; }}
  document.getElementById('stop-list-items').innerHTML = STOPS.slice(1).map((s,i) => {{
    const idx=i+1; const isDone=idx<curIdx; const isNow=idx===curIdx;
    return '<div class="stop-item'+(isDone?' done':'')+(isNow?' active-item':'')+'">'
      +'<span class="item-num'+(isDone?' done-num':'')+'">'+( isDone?'✓':idx)+'</span>'
      +'<span class="item-name">'+s.name+'</span>'
      +(isNow?'<span style="color:#f5a623">▶</span>':'')+'</div>';
  }}).join('');
}}
render();
</script>
</body></html>"""



@app.route("/api/reorder/<driver_name>", methods=["POST"])
def api_reorder(driver_name):
    """
    Reorder stops for a driver with optional locked positions.

    Body: { "stores": [ {name, lat, lng, locked: bool}, ... ] }

    锁定的门店通过 locked_indices 传入 optimize_route，被纳入完整 N×N 距离矩阵。
    OR-Tools 在感知所有门店地理位置的前提下，通过序列维度约束保持锁定门店的
    相对顺序，同时全局优化未锁定门店的插入位置。
    """
    if driver_name not in DRIVERS:
        return jsonify({"ok": False, "error": "Okänd chaufför"}), 404

    store_list = request.json.get("stores", [])
    if not store_list:
        return jsonify({"ok": False, "error": "Tom butikslista"}), 400

    # ── Patch missing coordinates from source file ──────────────
    # Happens when last_results.json was saved in old format without lat/lng.
    needs_patch = any(
        str(s.get("lat","")).strip() in ("","nan","None") or
        str(s.get("lng","")).strip() in ("","nan","None")
        for s in store_list
    )
    if needs_patch:
        coord_dict = load_coord_dict()
        patched = []
        for s in store_list:
            lat = str(s.get("lat","")).strip()
            lng = str(s.get("lng","")).strip()
            if lat in ("","nan","None") or lng in ("","nan","None"):
                key = s.get("name","").strip().lower()
                if key in coord_dict:
                    s = {**s,
                         "lat": str(coord_dict[key]["Latitude"]),
                         "lng": str(coord_dict[key]["Longitude"])}
                    print(f"[REORDER] Patched coords for '{s['name']}'")
                else:
                    print(f"[REORDER] WARNING: no coords found for '{s.get('name')}'")
            patched.append(s)
        store_list = patched

    # ── 收集锁定索引，剥离 locked 字段 ──────────────────────────
    # 锁定门店保持相对顺序的约束由 optimize_route / OR-Tools 处理，
    # 此处不再做任何手工拆分或拼合。
    locked_indices = {i for i, s in enumerate(store_list) if s.get("locked")}
    all_stores = [{k: v for k, v in s.items() if k != "locked"} for s in store_list]

    locked_names = [all_stores[i]["name"] for i in sorted(locked_indices)]
    print(f"[REORDER] {driver_name}: {len(all_stores)} 站 | "
          f"锁定 {len(locked_indices)} 站: {locked_names}")

    # ── 整体优化（含锁定约束）──────────────────────────────────
    optimization_warning = None
    locks_honored = True  # 默认认为锁定被满足（无锁定时也为 True）

    optimized, stats_or_err = optimize_route(
        all_stores,
        locked_indices=locked_indices if locked_indices else None,
    )

    if not optimized:
        # Fallback: keep current order
        print(f"[REORDER] optimize_route failed for {driver_name}: {stats_or_err}, using current order")
        optimized = all_stores
        optimization_warning = str(stats_or_err)
        locks_honored = bool(locked_indices)  # 原样保留 = 锁定自然成立
    else:
        if isinstance(stats_or_err, dict):
            locks_honored = stats_or_err.get("locks_honored", True)
            print(f"[REORDER] locks_honored={locks_honored}")

    final = optimized

    # ── 复用 optimize_route 返回的 stats，避免重复 OSRM 请求 ──
    # 仅在 optimize_route 失败（fallback 到原序）时才补充获取 stats。
    stats = None
    if isinstance(stats_or_err, dict) and "duration_sec" in stats_or_err:
        stats = stats_or_err
        print(f"[REORDER] 复用 optimize_route stats: "
              f"dur={stats.get('duration_sec')}s, dist={stats.get('distance_km')}km")
    else:
        print(f"[REORDER] Getting stats for final route ({len(final)} stores)…")
        stats = get_route_stats(final)
        if stats:
            print(f"[REORDER] get_route_stats: "
                  f"dur={stats.get('duration_sec')}s, dist={stats.get('distance_km')}km")
        else:
            print(f"[REORDER] get_route_stats returned None!")

    urls = generate_urls(final)

    # ── Update state ───────────────────────────────────────────
    r = state["results"].setdefault(driver_name, {})
    r["status"]        = "ok"
    r["stores"]        = [s["name"] for s in final]
    r["store_objects"] = final
    r["store_count"]   = len(final)
    r["urls"]          = urls
    if stats:
        dur_sec = stats.get("duration_sec", stats.get("duration_min", 0) * 60)
        hours   = dur_sec // 3600
        mins    = (dur_sec % 3600) // 60
        r["duration"]      = f"{hours} h {mins} min" if hours > 0 else f"{mins} min"
        r["duration_sec"]  = dur_sec
        r["distance"]      = f"{stats['distance_km']} km"
    elif not r.get("duration"):
        r["duration"] = "—"
        r["distance"] = "—"
    save_state()

    return jsonify({
        "ok":     True,
        "stores": final,
        "urls":   urls,
        "duration": r.get("duration"),
        "distance": r.get("distance"),
        "locks_honored": locks_honored,
        "warning": optimization_warning,
    })


@app.route("/api/emails", methods=["POST"])
def api_set_emails():
    for driver, addr in request.json.items():
        if driver in driver_emails:
            driver_emails[driver] = str(addr).strip()
    save_emails()
    return jsonify({"ok": True})


@app.route("/api/email-config", methods=["POST"])
def api_set_email_config():
    data = request.json
    if "sender" in data:
        email_config["sender"] = data["sender"]
    save_email_config()
    return jsonify({"ok": True})


@app.route("/api/email-config", methods=["GET"])
def api_get_email_config():
    return jsonify({
        "sender":  email_config.get("sender", ""),
        "api_key": "••••••" if email_config.get("api_key") else "",
    })


@app.route("/api/send-email/<driver_name>", methods=["POST"])
def api_send_email_one(driver_name):
    r = state["results"].get(driver_name)
    if not r or r.get("status") != "ok":
        return jsonify({"ok": False, "msg": "Inga rutter"}), 400
    base = request.host_url.rstrip("/")
    ok, msg = send_email_to_driver(driver_name, r, base)
    print(f"[EMAIL] {driver_name}: ok={ok} msg={msg}")
    return jsonify({"ok": ok, "msg": msg})


@app.route("/api/send-email-all", methods=["POST"])
def api_send_email_all():
    results_out = {}
    base = request.host_url.rstrip("/")
    for driver in DRIVERS:
        r = state["results"].get(driver)
        if r and r.get("status") == "ok":
            ok, msg = send_email_to_driver(driver, r, base)
            results_out[driver] = {"ok": ok, "msg": msg}
        else:
            results_out[driver] = {"ok": False, "msg": "Inga rutter"}
    return jsonify(results_out)

# ── 启动 ─────────────────────────────────────────────────────
# 启动时始终加载状态和调度器（gunicorn 也需要）
load_state()
scheduler.start()
reschedule(state["schedule_hour"], state["schedule_minute"])

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5050))
    app.run(debug=False, host="0.0.0.0", port=port, use_reloader=False)